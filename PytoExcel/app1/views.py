from django.shortcuts import render, HttpResponse
from app1.models import Generate_timesheet
from copy import deepcopy
import pandas as pd
import calendar
from openpyxl import Workbook
from datetime import date, datetime, timedelta
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import Alignment
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from rest_framework import viewsets
from serializers import UserSerializer
from sheet2dict import Worksheet
import excel2json


def user(request):
    with open('./sample.xlsx', "rb") as file:
        response = HttpResponse()
        response['Content-Disposition'] = 'attachment; filename=sample.xlsx'
        xlsx_data = generate_monthly_work_report(request)
        response.write(xlsx_data)
    return response


@csrf_exempt
def generate_monthly_work_report(request):
    if request.method == "POST":
        print("hey")
        employee_name = request.POST.get('employee_name')
        client_name = request.POST.get('client_name')
        project_manager = request.POST.get('project_manager')
        month_data = int(request.POST.get('month_data'))
        year_data = int(request.POST.get('year_data'))
        leave_taken = request.POST.get('leave_taken')
        holidays = request.POST.get('holidays')
        extra_work = request.POST.get('extra_workdays')
        print(employee_name)

        user = Generate_timesheet.objects.create(employee_name=employee_name,
                                                 client_name=client_name,
                                                 project_manager=project_manager, month_data=month_data,
                                                 year_data=year_data,
                                                 leave_taken=leave_taken, holidays=holidays,
                                                 extra_work=extra_work)

        book = Workbook()
        sheet = book.active

        sheet.column_dimensions['A'].width = 30
        sheet.column_dimensions['B'].width = 30
        sheet.column_dimensions['C'].width = 30
        sheet.column_dimensions['D'].width = 30

        weekly_holiday = ["Saturday", "Sunday"]
        first, last = calendar.monthrange(year_data, month_data)
        d1 = date(year=year_data, month=month_data, day=1)
        d2 = date(year=year_data, month=month_data, day=last) + timedelta(days=1)
        delta = d2 - d1

        sheet.merge_cells('A1:E1')
        sheet['A1'].fill = PatternFill("solid", start_color="FFFF00")

        cell = sheet.cell(row=1, column=1)
        cell.value = 'Timesheet of month of June 2022'
        cell.alignment = Alignment(horizontal='center', vertical='center')

        today = date.today()
        employee = ['employee name', employee_name, 'Client name', client_name]
        sheet.append(employee)

        sheet.merge_cells('A3:E3')
        sheet['A3'].fill = PatternFill("solid", start_color="FFA07A")
        cell = sheet.cell(row=3, column=1)
        cell.value = 'CH- Client Holiday, LT- Leave Taken, WH- Weekly Holiday'
        cell.alignment = Alignment(horizontal='center', vertical='center')
        row_data = {
            'Date': None,
            'Day': None,
            'Total_hours': None,
            'Particular': 'Python'
        }
        fieldnames = ['Date', 'Day', 'Total_hours', 'Particular']
        sheet.append(fieldnames)
        sheet_data = []
        count_wh = 0
        count_wd = 0
        count_hd = 0
        count_lt = 0
        for i in range(delta.days):
            a = d1 + timedelta(days=i)
            b = a.strftime('%A')
            if b in weekly_holiday:
                row_value = row_data.copy()
                row_value["Date"] = a
                row_value["Day"] = b
                row_value["Total_hours"] = 'WH'
                count_wh = count_wh + 1
            elif a is holidays:
                row_value = row_data.copy()
                row_value["Date"] = a
                row_value["Day"] = b
                row_value["Total_hours"] = "Holiday"
                count_hd = count_hd + 1
            elif a is leave_taken:
                row_value = row_data.copy()
                row_value["Date"] = a
                row_value["Day"] = b
                row_value["Total_hours"] = "LT"
                count_lt = count_lt + 1
            else:
                row_value = row_data.copy()
                row_value["Date"] = a
                row_value["Day"] = b
                row_value["Total_hours"] = 9
                count_wd = count_wd + 1
            sheet_data.append(row_value)

        for each_data in sheet_data:
            sheet.append(list(each_data.values()))

        sheet.merge_cells('A36:E36')
        sheet['A36'].fill = PatternFill("solid", start_color="DE3163")
        cell = sheet.cell(row=36, column=1)
        cell.value = 'Approval'

        work1 = ["Total No.of Days Worked", count_wd, "Name of the Project Manager", project_manager]
        sheet.append(work1)

        work2 = ["Total Leaves Taken", count_lt, "Signature of the Candidate", employee_name]
        sheet.append(work2)

        work3 = ["No. of other Holidays", count_hd]
        sheet.append(work3)

        fill_row = []
        fill_column = 0
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value in weekly_holiday:
                    fill_row.append(cell.row)
        fill_column = cell.column
        for each_row in fill_row:
            for each_column in range(1, fill_column):
                sheet.cell(row=each_row, column=each_column).fill = PatternFill("solid", start_color="424242")

        rows = range(1, 44)
        columns = range(1, 10)
        for row in rows:
            for col in columns:
                sheet.cell(row, col).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        book.save("sample.xlsx")

    with open('./sample.xlsx', "rb") as file:
        response = HttpResponse(file, content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="sample.xlsx"'
    return response

    #     return render(request, "data.html", context={"user": user})
    # return render(request, "data.html")


class UserViewSet(viewsets.ModelViewSet):
    """
    API endpoint that allows users to be viewed or edited.
    """
    queryset = Generate_timesheet.objects.all()
    serializer_class = UserSerializer

# def generate_monthly_work_report(*args):
#     print("hey")
#     employee_name = input("Enter Employee Name: ")
#     client_name = input("Enter Client: ")
#     project_manager = input("Enter name of Project Manager: ")
#     month_data = int(input("Enter Month in format (01-12): "))
#     year_data = int(input("Enter Year: "))
#     first, last = calendar.monthrange(year_data, month_data)
#     print(first, last)
#     leave_taken = []
#     m = int(input("Enter number of Leave taken in the month: "))
#     for i in range(0, m):
#         dates_lt = input("Enter dates of Leaves taken in format (yy-mm-dd):")
#         leave_taken.append(datetime.strptime(dates_lt, "%Y-%m-%d").date())
#     holidays = []
#     x = int(input("Enter number of holidays in the month: "))
#     for i in range(0, x):
#         dates_hd = input("Enter dates of Holidays in format (yy-mm-dd):")
#         holidays.append(datetime.strptime(dates_hd, "%Y-%m-%d").date())
#     extra_work = []
#     n = int(input("Enter number of extra days work: "))
#     for i in range(0, n):
#         dates_ew = input("Enter dates of extra work in format (yy-mm-dd):")
#         extra_work.append(datetime.strptime(dates_ew, "%Y-%m-%d").date())
#     book = Workbook()
#     sheet = book.active
#
#     sheet.column_dimensions['A'].width = 30
#     sheet.column_dimensions['B'].width = 30
#     sheet.column_dimensions['C'].width = 30
#     sheet.column_dimensions['D'].width = 30
#
#     weekly_holiday = ["Saturday", "Sunday"]
#
#     d1 = date(year=year_data, month=month_data, day=1)
#     d2 = date(year=year_data, month=month_data, day=last) + timedelta(days=1)
#     delta = d2 - d1
#
#     sheet.merge_cells('A1:E1')
#     sheet['A1'].fill = PatternFill("solid", start_color="FFFF00")
#
#     cell = sheet.cell(row=1, column=1)
#     cell.value = 'Timesheet of month of June 2022'
#     cell.alignment = Alignment(horizontal='center', vertical='center')
#
#     today = date.today()
#     employee = ['employee name', employee_name, 'Client name', client_name]
#     sheet.append(employee)
#
#     sheet.merge_cells('A3:E3')
#     sheet['A3'].fill = PatternFill("solid", start_color="FFA07A")
#     cell = sheet.cell(row=3, column=1)
#     cell.value = 'CH- Client Holiday, LT- Leave Taken, WH- Weekly Holiday'
#     cell.alignment = Alignment(horizontal='center', vertical='center')
#     row_data = {
#         'Date': None,
#         'Day': None,
#         'Total hours': None,
#         'Particular': 'Python'
#     }
#
#     fieldnames = ['Date', 'Day', 'Total hours', 'Particular']
#     sheet.append(fieldnames)
#     sheet_data = []
#     count_wh = 0
#     count_wd = 0
#     count_hd = 0
#     count_lt = 0
#     for i in range(delta.days):
#         a = d1 + timedelta(days=i)
#         b = a.strftime('%A')
#         if b in weekly_holiday:
#             row_value = row_data.copy()
#             row_value["Date"] = a
#             row_value["Day"] = b
#             row_value["Total hours"] = 'WH'
#             count_wh = count_wh + 1
#         elif a in holidays:
#             row_value = row_data.copy()
#             row_value["Date"] = a
#             row_value["Day"] = b
#             row_value["Total hours"] = "Holiday"
#             count_hd = count_hd + 1
#         elif a in leave_taken:
#             row_value = row_data.copy()
#             row_value["Date"] = a
#             row_value["Day"] = b
#             row_value["Total hours"] = "LT"
#             count_lt = count_lt + 1
#         else:
#             row_value = row_data.copy()
#             row_value["Date"] = a
#             row_value["Day"] = b
#             row_value["Total hours"] = 9
#             count_wd = count_wd + 1
#         sheet_data.append(row_value)
#
#     for each_data in sheet_data:
#         sheet.append(list(each_data.values()))
#
#     sheet.merge_cells('A36:E36')
#     sheet['A36'].fill = PatternFill("solid", start_color="DE3163")
#     cell = sheet.cell(row=36, column=1)
#     cell.value = 'Approval'
#
#     work1 = ["Total No.of Days Worked", count_wd, "Name of the Project Manager", project_manager]
#     sheet.append(work1)
#
#     work2 = ["Total Leaves Taken", count_lt, "Signature of the Candidate", employee_name]
#     sheet.append(work2)
#
#     work3 = ["No. of other Holidays", count_hd]
#     sheet.append(work3)
#
#     fill_row = []
#     fill_column = 0
#     for row in sheet.iter_rows():
#         for cell in row:
#             if cell.value in weekly_holiday:
#                 fill_row.append(cell.row)
#     fill_column = cell.column
#     for each_row in fill_row:
#         for each_column in range(1, fill_column):
#             sheet.cell(row=each_row, column=each_column).fill = PatternFill("solid", start_color="424242")
#
#     rows = range(1, 44)
#     columns = range(1, 10)
#     for row in rows:
#         for col in columns:
#             sheet.cell(row, col).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
#     book.save("sample.xlsx")
#     return "sample.xlsx"
#
# generate_monthly_work_report()


# df = pd.read_excel("sample.xlsx", skiprows=[0, 1, 2])
# key = df.columns
# print(key)
#
# person = {}
# dates = list(df.Date)
# days = list(df.Day)
# total_hours = list(df.Total_hours)
# particular = list(df.Particular)
#
# for data in dates:
#     for day in days:
#         for total in total_hours:
#             for part in particular:
#                 person = {"dates": dates, "days": days, "total_hours": total_hours,
#                         "particular": particular}
# print(person)
#
# return JsonResponse(person)
